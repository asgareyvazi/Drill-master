"""Input adapters for the Universal Import entry point.

Professional Features:
- CSV to XLSX with encoding detection (utf-8-sig, utf-8, cp1256 for Persian)
- PDF to XLSX with 3-tier extraction:
  1. Camelot (text PDF tables) - preferred
  2. PyMuPDF coordinate-preserving text extraction - fallback
  3. OCR via pytesseract + pdf2image for scanned PDFs - final fallback
- Metrics preserved for Review Matrix and Data Quality Score
- Atomic transaction ready
"""

from pathlib import Path
import csv
import re
import logging
from typing import Dict, Any, List

logger = logging.getLogger(__name__)


def _write_rows(rows, output_path, sheet_name="Imported"):
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or "Imported"
    for row in rows:
        sheet.append(list(row))
    workbook.save(output_path)
    workbook.close()
    return {"rows": len(rows), "path": str(output_path)}


def csv_to_xlsx(csv_path, output_path):
    """CSV to XLSX with intelligent multi-section detection.
    
    DDR CSV files typically have multiple sections separated by blank rows:
    - Well Info header
    - Time Log table
    - Mud Properties table
    - Drilling Parameters table
    - Survey table
    
    This function detects section boundaries and creates separate sheets
    when possible, or at minimum preserves the structure for the scanner.
    """
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "windows-1256", "iso-8859-6"):
        try:
            with open(csv_path, newline="", encoding=encoding) as stream:
                reader = csv.reader(stream)
                rows = list(reader)
                if not rows:
                    continue
                
                # Detect section boundaries (blank rows or keyword rows)
                sections = _detect_csv_sections(rows)
                
                if len(sections) > 1:
                    # Multi-section CSV → multiple sheets
                    result = _write_multi_sheet(rows, sections, output_path, Path(csv_path).stem)
                else:
                    # Single section → one sheet
                    result = _write_rows(rows, output_path, Path(csv_path).stem)
                
                result["encoding"] = encoding
                result["engine"] = "csv"
                result["sections_detected"] = len(sections)
                return result
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            logger.debug(f"CSV read with {encoding} failed: {exc}")
            continue
    raise ValueError("CSV encoding is not supported - tried utf-8-sig, utf-8, cp1256, windows-1256")


# Section keywords for DDR detection
_SECTION_KEYWORDS = {
    "time_log": ["time log", "24h", "24 hour", "activity", "from", "to", "duration", "main code"],
    "mud": ["mud", "weight", "pv", "yp", "gel", "funnel", "ph", "rheology", "density"],
    "drilling": ["drilling", "bit", "wob", "rpm", "torque", "rop", "pump", "nozzle"],
    "survey": ["survey", "md", "inc", "azi", "tvd", "inclination", "azimuth", "measured depth"],
    "bha": ["bha", "bottom hole", "component", "stabilizer", "drill collar"],
    "safety": ["safety", "hse", "lti", "incident", "bop", "drill"],
    "logistics": ["pob", "personnel", "fuel", "water", "bulk", "transport"],
    "casing": ["casing", "size", "weight", "grade", "depth"],
    "cost": ["cost", "afe", "expense", "budget"],
}


def _detect_csv_sections(rows: list) -> list:
    """Detect section boundaries in a flat CSV.
    
    Returns list of (start_row, end_row, section_name) tuples.
    """
    sections = []
    current_start = 0
    current_type = "header"
    blank_count = 0
    
    for i, row in enumerate(rows):
        # Check if row is blank
        is_blank = all(cell.strip() == "" for cell in row if cell)
        
        if is_blank:
            blank_count += 1
            if blank_count >= 2 and current_start < i:
                # End of section
                sections.append((current_start, i, current_type))
                current_start = i + 1
                current_type = "unknown"
                blank_count = 0
        else:
            blank_count = 0
            # Try to detect section type from first non-blank row after boundary
            if current_type == "unknown":
                row_text = " ".join(cell.lower().strip() for cell in row if cell)
                for section_type, keywords in _SECTION_KEYWORDS.items():
                    if any(kw in row_text for kw in keywords):
                        current_type = section_type
                        break
    
    # Last section
    if current_start < len(rows):
        sections.append((current_start, len(rows), current_type))
    
    return sections


def _write_multi_sheet(rows: list, sections: list, output_path: str, base_name: str) -> dict:
    """Write multiple sections to separate sheets in an XLSX file."""
    from openpyxl import Workbook
    
    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)
    
    total_rows = 0
    for start, end, section_type in sections:
        section_rows = rows[start:end]
        # Filter out completely blank rows
        section_rows = [r for r in section_rows if any(cell.strip() for cell in r if cell)]
        
        if not section_rows:
            continue
        
        sheet_name = f"{section_type[:20]}_{start}" if section_type != "unknown" else f"Section_{start}"
        sheet = workbook.create_sheet(title=sheet_name[:31])
        
        for row in section_rows:
            sheet.append(list(row))
        total_rows += len(section_rows)
    
    if not workbook.worksheets:
        # Fallback: single sheet with all rows
        sheet = workbook.create_sheet(title=base_name[:31] or "Imported")
        for row in rows:
            sheet.append(list(row))
        total_rows = len(rows)
    
    workbook.save(output_path)
    workbook.close()
    
    return {"rows": total_rows, "path": str(output_path), "sheets": len(workbook.worksheets)}


def pdf_to_xlsx(pdf_path, output_path):
    """Convert PDF tables to XLSX with 3-tier extraction.

    Tier 1: Camelot (text PDF tables) - best for bordered tables
    Tier 2: PyMuPDF coordinate-preserving text - good for text PDFs without borders
    Tier 3: OCR via pytesseract - for scanned PDFs (Qwen-VL future, pytesseract now)

    Returned metrics are kept for Import Review Matrix and Data Quality.
    """
    metrics: List[Dict[str, Any]] = []
    rows: List[List[str]] = []

    # Tier 1: Camelot
    try:
        import camelot
        tables = camelot.read_pdf(str(pdf_path), pages="all", flavor="stream")
        tier1_rows = []
        for index, table in enumerate(tables, 1):
            report = getattr(table, "parsing_report", {})
            metrics.append({"tier": 1, "table": index, "engine": "camelot", **report})
            tier1_rows.append([f"[Table {index} | Page {report.get('page', '')} | Accuracy {report.get('accuracy', '')}]"])
            tier1_rows.extend(table.df.fillna("").astype(str).values.tolist())
        if tier1_rows:
            result = _write_rows(tier1_rows, output_path, "PDF Tables - Camelot")
            result["engine"] = "camelot"
            result["tier"] = 1
            result["metrics"] = metrics
            result["table_count"] = len(tables)
            logger.info(f"PDF Tier1 Camelot: {len(tables)} tables from {pdf_path}")
            return result
        else:
            metrics.append({"tier": 1, "engine": "camelot", "info": "No tables found"})
    except ImportError:
        metrics.append({"tier": 1, "engine": "camelot", "error": "not installed - pip install camelot-py[cv]"})
    except Exception as exc:
        metrics.append({"tier": 1, "engine": "camelot", "error": str(exc)})
        logger.debug(f"Camelot failed: {exc}")

    # Tier 2: PyMuPDF text extraction with coordinate grouping
    try:
        import fitz  # PyMuPDF

        document = fitz.open(str(pdf_path))
        tier2_rows = []
        try:
            for page_no, page in enumerate(document, 1):
                blocks = page.get_text("blocks")
                grouped: Dict[float, List[tuple]] = {}
                for block in blocks:
                    x0, y0, _x1, _y1, text = block[:5]
                    for line in str(text).splitlines():
                        line = line.strip()
                        if line:
                            grouped.setdefault(round(y0, 1), []).append((x0, line))

                tier2_rows.append([f"[Page {page_no} - PyMuPDF Text]"])
                for _y, cells in sorted(grouped.items()):
                    cells.sort(key=lambda value: value[0])
                    text = " | ".join(value for _x, value in cells)
                    parts = [part.strip() for part in re.split(r"\s*\|\s*|\t|\s{2,}", text) if part.strip()]
                    if parts:
                        tier2_rows.append(parts)

            if tier2_rows and len(tier2_rows) > len(document):  # More than just page headers
                result = _write_rows(tier2_rows, output_path, "PDF Text - PyMuPDF")
                result["engine"] = "pymupdf-coordinate-fallback"
                result["tier"] = 2
                result["metrics"] = metrics
                result["page_count"] = len(document)
                logger.info(f"PDF Tier2 PyMuPDF: {len(tier2_rows)} rows from {pdf_path}")
                return result
            else:
                metrics.append({"tier": 2, "engine": "pymupdf", "info": "No meaningful text extracted, trying OCR"})

        finally:
            document.close()

    except ImportError as exc:
        metrics.append({"tier": 2, "engine": "pymupdf", "error": f"not installed: {exc}"})
    except Exception as exc:
        metrics.append({"tier": 2, "engine": "pymupdf", "error": str(exc)})
        logger.debug(f"PyMuPDF failed: {exc}")

    # Tier 3: OCR for scanned PDFs
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image

        # Convert PDF pages to images
        images = convert_from_path(str(pdf_path), first_page=1, last_page=5)  # limit to 5 pages for performance
        tier3_rows = []

        for page_no, image in enumerate(images, 1):
            tier3_rows.append([f"[Page {page_no} - OCR]"])
            # OCR with table awareness
            ocr_text = pytesseract.image_to_string(image)

            for line in ocr_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                # Split by multiple spaces or tabs as table columns
                parts = [p.strip() for p in re.split(r"\t|\s{2,}| \| ", line) if p.strip()]
                if parts:
                    tier3_rows.append(parts)

            # Also try to get data with TSV for better structure
            try:
                tsv_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DATAFRAME)
                # Group by block/line
                if not tsv_data.empty:
                    # Filter confident text
                    confident = tsv_data[tsv_data.conf > 60]
                    if not confident.empty:
                        metrics.append(
                            {
                                "tier": 3,
                                "page": page_no,
                                "engine": "pytesseract",
                                "confident_words": len(confident),
                                "avg_conf": float(confident.conf.mean()),
                            }
                        )
            except Exception:
                pass

        if tier3_rows and len(tier3_rows) > len(images):
            result = _write_rows(tier3_rows, output_path, "PDF OCR - Tesseract")
            result["engine"] = "pytesseract-ocr"
            result["tier"] = 3
            result["metrics"] = metrics
            result["page_count"] = len(images)
            logger.info(f"PDF Tier3 OCR: {len(tier3_rows)} rows from {pdf_path}")
            return result
        else:
            metrics.append({"tier": 3, "engine": "pytesseract", "info": "OCR produced no data"})

    except ImportError as exc:
        metrics.append({"tier": 3, "engine": "pytesseract", "error": f"not installed: {exc} - pip install pytesseract pdf2image Pillow"})
    except Exception as exc:
        metrics.append({"tier": 3, "engine": "pytesseract", "error": str(exc)})
        logger.debug(f"OCR failed: {exc}")

    # If all tiers failed but we have some rows from any tier, return what we have
    if rows:
        result = _write_rows(rows, output_path, "PDF Fallback")
        result["engine"] = "fallback"
        result["metrics"] = metrics
        return result

    # Complete failure - return metrics for Review Matrix
    raise RuntimeError(
        f"PDF import failed for {pdf_path} - All 3 tiers failed:\n"
        f"Tier1 Camelot: {metrics[0] if len(metrics)>0 else 'no attempt'}\n"
        f"Tier2 PyMuPDF: {metrics[1] if len(metrics)>1 else 'no attempt'}\n"
        f"Tier3 OCR: {metrics[2] if len(metrics)>2 else 'no attempt'}\n"
        f"Install: pip install camelot-py[cv] PyMuPDF pytesseract pdf2image Pillow"
    )


def extract_pdf_with_ocr_preview(pdf_path: str) -> Dict[str, Any]:
    """Extract PDF with preview of which tier will be used, for Import Preview dialog.

    Returns dict with tier, engine, table count, sample rows, metrics.
    No file writing - just analysis.
    """
    metrics = []
    preview = {"file": str(pdf_path), "tier": None, "engine": None, "table_count": 0, "sample_rows": [], "metrics": []}

    # Tier1 check
    try:
        import camelot
        tables = camelot.read_pdf(str(pdf_path), pages="1", flavor="stream")
        if tables and len(tables) > 0:
            preview["tier"] = 1
            preview["engine"] = "camelot"
            preview["table_count"] = len(tables)
            preview["sample_rows"] = tables[0].df.head(3).astype(str).values.tolist() if hasattr(tables[0], 'df') else []
            preview["metrics"] = [{"tier": 1, "engine": "camelot", "tables": len(tables)}]
            return preview
    except Exception as exc:
        metrics.append({"tier": 1, "error": str(exc)})

    # Tier2 check
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        text = doc[0].get_text() if len(doc) > 0 else ""
        doc.close()
        if text and len(text.strip()) > 100:
            preview["tier"] = 2
            preview["engine"] = "pymupdf"
            preview["sample_rows"] = [line.split() for line in text.splitlines()[:3] if line.strip()]
            preview["metrics"] = metrics + [{"tier": 2, "engine": "pymupdf", "text_length": len(text)}]
            return preview
    except Exception as exc:
        metrics.append({"tier": 2, "error": str(exc)})

    # Tier3 - scanned
    preview["tier"] = 3
    preview["engine"] = "pytesseract (scanned PDF suspected)"
    preview["metrics"] = metrics + [{"tier": 3, "info": "Scanned PDF - OCR required"}]
    return preview
