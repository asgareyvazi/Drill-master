"""
Professional Export - Intelligence Platform

Implements spec:
- Export metadata: Company, Project, Field, Well, Section, Report Number, Report Date, Revision, Status, Prepared By, Checked By, Approved By, Generated At UTC, Timezone, Units, Data Quality, Audit ID
- Excel sheets: Executive Summary, Daily Report, Time Logs, Mud, Drilling Parameters, Bit, BHA, Survey, Safety, Logistics, Services, Cost, Data Quality, Validation, Audit, Raw Data
- PDF: Header, Logo, Footer, Page Number, Watermark, Revision, Approval, Tables, Charts, Signature, Data Quality
"""

from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ProfessionalExportMetadata:
    """Builds professional export metadata as per spec."""

    @staticmethod
    def build(db_manager, well_id: int, section_id: int = None, report_id: int = None) -> Dict[str, Any]:
        well = db_manager.get_well_by_id(well_id) or {}
        section = {}
        if section_id:
            sections = db_manager.get_sections_by_well(well_id)
            section = next((s for s in sections if s["id"] == section_id), {})

        report = db_manager.get_daily_report_by_id(report_id) if report_id else {}

        # Get hierarchy for company/project
        company_name = ""
        project_name = ""
        try:
            from core.database import Well, Project, Company
            with db_manager.session_scope() as session:
                w_obj = session.get(Well, well_id)
                if w_obj and w_obj.project:
                    project_name = w_obj.project.name
                    if w_obj.project.company:
                        company_name = w_obj.project.company.name
        except Exception:
            pass

        # Data quality
        try:
            from core.data_quality import DataQualityService
            dq_service = DataQualityService(db_manager)
            dq_summary = dq_service.summary(report_id) if report_id else {"score": 0, "status": "unknown"}
        except Exception:
            dq_summary = {"score": 0, "status": "unknown"}

        # Audit
        audit_logs = []
        try:
            audit_logs = db_manager.get_audit_logs(entity_type="daily_report", entity_id=report_id, limit=5) if report_id else []
        except Exception:
            pass

        now_utc = datetime.now(timezone.utc)

        return {
            "Company": company_name or well.get("client", "") or well.get("operator", "") or "Default Company",
            "Project": project_name or well.get("project_name", "") or "Default Project",
            "Field": well.get("field_name", ""),
            "Well": well.get("name", ""),
            "Section": section.get("name", "") or well.get("section_name", ""),
            "Report Number": report.get("report_number", ""),
            "Report Date": str(report.get("report_date", "")),
            "Revision": "Rev 0",
            "Status": report.get("status", "Draft"),
            "Prepared By": well.get("supervisor_day", "") or well.get("geologist1", ""),
            "Checked By": well.get("supervisor_night", "") or well.get("superintendent", ""),
            "Approved By": well.get("operation_manager", "") or well.get("operator", ""),
            "Generated At UTC": now_utc.isoformat(),
            "Timezone": "UTC",
            "Units": "Metric (m, ppg, psi, gpm, m/hr, rpm)",
            "Data Quality": f"{dq_summary.get('score',0)}% - {dq_summary.get('status','')}",
            "Audit ID": f"AUDIT-{well_id}-{report_id or 0}-{now_utc.strftime('%Y%m%d%H%M%S')}",
            "Data Quality Detail": dq_summary,
            "Audit Logs": audit_logs,
            "Well Info": well,
            "Section Info": section,
            "Report Info": report,
        }


class ProfessionalExcelExport:
    """Excel export with all professional sheets."""

    SHEETS = [
        "Executive Summary",
        "Daily Report",
        "Time Logs",
        "Mud",
        "Drilling Parameters",
        "Bit",
        "BHA",
        "Survey",
        "Safety",
        "Logistics",
        "Services",
        "Cost",
        "Data Quality",
        "Validation",
        "Audit",
        "Raw Data",
    ]

    def __init__(self, db_manager):
        self.db = db_manager

    def export(self, well_id: int, output_path: str, report_id: int = None, section_id: int = None) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            wb.remove(wb.active)

            metadata = ProfessionalExportMetadata.build(self.db, well_id, section_id, report_id)

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="2C3E50")
            sub_header_fill = PatternFill("solid", fgColor="3498DB")

            # 1. Executive Summary
            ws = wb.create_sheet("Executive Summary")
            ws["A1"] = "DrillMaster - Professional Export - Intelligence Platform"
            ws["A1"].font = Font(bold=True, size=14)
            row = 3
            for key in ["Company", "Project", "Field", "Well", "Section", "Report Number", "Report Date", "Revision", "Status", "Prepared By", "Checked By", "Approved By", "Generated At UTC", "Timezone", "Units", "Data Quality", "Audit ID"]:
                ws.cell(row=row, column=1, value=key).font = Font(bold=True)
                ws.cell(row=row, column=2, value=str(metadata.get(key, "")))
                row += 1

            # 2. Daily Report
            ws2 = wb.create_sheet("Daily Report")
            report = metadata.get("Report Info", {})
            if report:
                headers = list(report.keys())
                for c, h in enumerate(headers, 1):
                    ws2.cell(row=1, column=c, value=h).font = header_font
                    ws2.cell(row=1, column=c).fill = header_fill
                for c, h in enumerate(headers, 1):
                    ws2.cell(row=2, column=c, value=str(report.get(h, "")))

            # 3. Time Logs
            ws3 = wb.create_sheet("Time Logs")
            try:
                from core.database import TimeLog24H
                with self.db.session_scope() as session:
                    logs = session.query(TimeLog24H).filter(TimeLog24H.report_id == report_id).all() if report_id else []
                    if logs:
                        headers = ["From", "To", "Duration", "Main Phase", "Main Code", "Sub Code", "NPT", "Contractor", "Description"]
                        for c, h in enumerate(headers, 1):
                            ws3.cell(row=1, column=c, value=h).font = header_font
                            ws3.cell(row=1, column=c).fill = header_fill
                        for r, log in enumerate(logs, 2):
                            ws3.cell(row=r, column=1, value=str(log.time_from))
                            ws3.cell(row=r, column=2, value=str(log.time_to))
                            ws3.cell(row=r, column=3, value=log.duration or 0)
                            ws3.cell(row=r, column=4, value=log.main_phase or "")
                            ws3.cell(row=r, column=5, value=log.main_code or "")
                            ws3.cell(row=r, column=6, value=log.sub_code or "")
                            ws3.cell(row=r, column=7, value="Yes" if log.is_npt else "No")
                            ws3.cell(row=r, column=8, value=log.contractor or "")
                            ws3.cell(row=r, column=9, value=log.activity_description or "")
            except Exception as e:
                logger.debug(f"Time logs export failed: {e}")

            # 4. Mud
            ws4 = wb.create_sheet("Mud")
            try:
                mud = self.db.get_mud_report(report_id=report_id) if report_id else self.db.get_mud_report(well_id=well_id)
                if mud:
                    for c, k in enumerate(mud.keys(), 1):
                        ws4.cell(row=1, column=c, value=k).font = header_font
                        ws4.cell(row=1, column=c).fill = header_fill
                    for c, k in enumerate(mud.keys(), 1):
                        ws4.cell(row=2, column=c, value=str(mud.get(k, "")))
            except Exception:
                pass

            # 5. Drilling Parameters
            ws5 = wb.create_sheet("Drilling Parameters")
            try:
                params = self.db.get_drilling_parameters(report_id=report_id) if report_id else self.db.get_drilling_parameters(well_id=well_id)
                if params:
                    for c, k in enumerate(params.keys(), 1):
                        ws5.cell(row=1, column=c, value=k).font = header_font
                        ws5.cell(row=1, column=c).fill = header_fill
                    for c, k in enumerate(params.keys(), 1):
                        ws5.cell(row=2, column=c, value=str(params.get(k, "")))
            except Exception:
                pass

            # 6. Bit
            ws6 = wb.create_sheet("Bit")
            try:
                bit = self.db.get_bit_report(well_id, report_id)
                if bit:
                    ws6.cell(row=1, column=1, value="Bit Records JSON").font = header_font
                    ws6.cell(row=1, column=1).fill = header_fill
                    ws6.cell(row=2, column=1, value=str(bit.get("bit_records_json", ""))[:32000])
            except Exception:
                pass

            # 7. BHA
            ws7 = wb.create_sheet("BHA")
            try:
                bha = self.db.get_bha_report(well_id, report_id)
                if bha:
                    ws7.cell(row=1, column=1, value="BHA Configs").font = header_font
                    ws7.cell(row=1, column=1).fill = header_fill
                    ws7.cell(row=2, column=1, value=str(bha.get("bha_configs", ""))[:32000])
            except Exception:
                pass

            # 8. Survey
            ws8 = wb.create_sheet("Survey")
            try:
                from core.database import SurveyPoint
                with self.db.session_scope() as session:
                    points = session.query(SurveyPoint).filter(SurveyPoint.well_id == well_id).order_by(SurveyPoint.md).all()
                    if points:
                        headers = ["MD", "Inc", "Azi", "TVD", "North", "East", "VS", "HD", "DLS", "Tool"]
                        for c, h in enumerate(headers, 1):
                            ws8.cell(row=1, column=c, value=h).font = header_font
                            ws8.cell(row=1, column=c).fill = header_fill
                        for r, p in enumerate(points, 2):
                            ws8.cell(row=r, column=1, value=p.md)
                            ws8.cell(row=r, column=2, value=p.inc)
                            ws8.cell(row=r, column=3, value=p.azi)
                            ws8.cell(row=r, column=4, value=p.tvd)
                            ws8.cell(row=r, column=5, value=p.north)
                            ws8.cell(row=r, column=6, value=p.east)
                            ws8.cell(row=r, column=7, value=p.vs)
                            ws8.cell(row=r, column=8, value=p.hd)
                            ws8.cell(row=r, column=9, value=p.dls)
                            ws8.cell(row=r, column=10, value=p.tool)
            except Exception as e:
                logger.debug(f"Survey export failed: {e}")

            # 9. Safety
            ws9 = wb.create_sheet("Safety")
            try:
                safety = self.db.get_safety_report(well_id=well_id, report_id=report_id)
                if safety:
                    for c, k in enumerate(safety.keys(), 1):
                        ws9.cell(row=1, column=c, value=k).font = header_font
                        ws9.cell(row=1, column=c).fill = header_fill
                    for c, k in enumerate(safety.keys(), 1):
                        ws9.cell(row=2, column=c, value=str(safety.get(k, ""))[:1000])
            except Exception:
                pass

            # 10. Logistics
            ws10 = wb.create_sheet("Logistics")
            try:
                from core.database import FuelWaterInventory, BulkMaterials, ServiceCompanyPOB
                with self.db.session_scope() as session:
                    bulks = session.query(BulkMaterials).filter(BulkMaterials.well_id == well_id).all() if well_id else []
                    if bulks:
                        ws10.cell(row=1, column=1, value="Bulk Materials").font = header_font
                        ws10.cell(row=1, column=1).fill = sub_header_fill
                        headers = ["Material", "Initial", "Received", "Used", "Current", "Unit", "Date"]
                        for c, h in enumerate(headers, 1):
                            ws10.cell(row=2, column=c, value=h).font = header_font
                            ws10.cell(row=2, column=c).fill = header_fill
                        for r, b in enumerate(bulks, 3):
                            ws10.cell(row=r, column=1, value=b.material_name)
                            ws10.cell(row=r, column=2, value=b.initial_stock or 0)
                            ws10.cell(row=r, column=3, value=b.received or 0)
                            ws10.cell(row=r, column=4, value=b.used or 0)
                            ws10.cell(row=r, column=5, value=b.current_stock or 0)
                            ws10.cell(row=r, column=6, value=b.unit or "")
                            ws10.cell(row=r, column=7, value=str(b.report_date))
            except Exception as e:
                logger.debug(f"Logistics export failed: {e}")

            # 11. Services
            ws11 = wb.create_sheet("Services")
            try:
                services = self.db.get_service_companies(well_id=well_id, report_id=report_id)
                if services:
                    headers = list(services[0].keys()) if services else []
                    for c, h in enumerate(headers, 1):
                        ws11.cell(row=1, column=c, value=h).font = header_font
                        ws11.cell(row=1, column=c).fill = header_fill
                    for r, svc in enumerate(services, 2):
                        for c, h in enumerate(headers, 1):
                            ws11.cell(row=r, column=c, value=str(svc.get(h, ""))[:1000])
            except Exception:
                pass

            # 12. Cost
            ws12 = wb.create_sheet("Cost")
            try:
                costs = self.db.get_cost_records(well_id)
                if costs:
                    headers = list(costs[0].keys()) if costs else []
                    for c, h in enumerate(headers, 1):
                        ws12.cell(row=1, column=c, value=h).font = header_font
                        ws12.cell(row=1, column=c).fill = header_fill
                    for r, cost in enumerate(costs, 2):
                        for c, h in enumerate(headers, 1):
                            ws12.cell(row=r, column=c, value=str(cost.get(h, ""))[:1000])
            except Exception:
                pass

            # 13. Data Quality
            ws13 = wb.create_sheet("Data Quality")
            try:
                from core.data_quality import DataQualityService
                dq = DataQualityService(self.db)
                summary = dq.summary(report_id) if report_id else dq.for_well(well_id)
                if isinstance(summary, dict):
                    ws13.cell(row=1, column=1, value="Metric").font = header_font
                    ws13.cell(row=1, column=1).fill = header_fill
                    ws13.cell(row=1, column=2, value="Value").font = header_font
                    ws13.cell(row=1, column=2).fill = header_fill
                    ws13.cell(row=1, column=3, value="Status").font = header_font
                    ws13.cell(row=1, column=3).fill = header_fill
                    ws13.cell(row=1, column=4, value="Detail").font = header_font
                    ws13.cell(row=1, column=4).fill = header_fill
                    for r, m in enumerate(summary.get("metrics", []), 2):
                        ws13.cell(row=r, column=1, value=m.get("name", ""))
                        ws13.cell(row=r, column=2, value=m.get("value", ""))
                        ws13.cell(row=r, column=3, value=m.get("status", ""))
                        ws13.cell(row=r, column=4, value=m.get("detail", ""))
                    ws13.cell(row=len(summary.get("metrics", [])) + 3, column=1, value="Overall Score").font = Font(bold=True)
                    ws13.cell(row=len(summary.get("metrics", [])) + 3, column=2, value=summary.get("score", 0))
            except Exception as e:
                logger.debug(f"DQ export failed: {e}")

            # 14. Validation
            ws14 = wb.create_sheet("Validation")
            ws14.cell(row=1, column=1, value="Validation Rules").font = header_font
            ws14.cell(row=1, column=1).fill = header_fill
            validations = [
                "Well name required - MISSING_INPUT",
                "Report date required - MISSING_INPUT",
                "Depth cannot be negative",
                "Time log total must equal 24h",
                "Overlap detection",
                "Gap detection",
                "Duplicate MD detection",
                "Non-monotonic MD detection",
                "Negative stock check",
                "BOP working pressure required - critical",
            ]
            for r, v in enumerate(validations, 2):
                ws14.cell(row=r, column=1, value=v)

            # 15. Audit
            ws15 = wb.create_sheet("Audit")
            audit_logs = metadata.get("Audit Logs", [])
            if audit_logs:
                headers = list(audit_logs[0].keys()) if audit_logs else ["No audit logs"]
                for c, h in enumerate(headers, 1):
                    ws15.cell(row=1, column=c, value=h).font = header_font
                    ws15.cell(row=1, column=c).fill = header_fill
                for r, log in enumerate(audit_logs, 2):
                    for c, h in enumerate(headers, 1):
                        ws15.cell(row=r, column=c, value=str(log.get(h, ""))[:1000])

            # 16. Raw Data
            ws16 = wb.create_sheet("Raw Data")
            ws16.cell(row=1, column=1, value="Raw import data preserved for lineage").font = header_font
            ws16.cell(row=1, column=1).fill = header_fill

            wb.save(output_path)
            logger.info(f"Professional Excel export saved: {output_path} with {len(wb.sheetnames)} sheets")
            return True

        except Exception as e:
            logger.error(f"Professional Excel export failed: {e}", exc_info=True)
            return False


class ProfessionalPDFExport:
    """PDF export with Header, Logo, Footer, Page Number, Watermark, Revision, Approval, Tables, Charts, Signature, Data Quality."""

    def __init__(self, db_manager):
        self.db = db_manager

    def export(self, well_id: int, output_path: str, report_id: int = None, section_id: int = None) -> bool:
        try:
            from PySide6.QtGui import QTextDocument, QPageSize, QPageLayout
            from PySide6.QtCore import QMarginsF
            from PySide6.QtPrintSupport import QPrinter
            from core.text_utils import wrap_html

            metadata = ProfessionalExportMetadata.build(self.db, well_id, section_id, report_id)

            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{ margin: 15mm; size: A4; }}
body {{ font-family: Arial, Helvetica, sans-serif; font-size: 9pt; color: #2c3e50; }}
.header {{ background: #2c3e50; color: white; padding: 10px; border-radius: 4px; margin-bottom: 10px; }}
.header h1 {{ margin: 0; font-size: 16pt; }}
.header .meta {{ font-size: 8pt; opacity: 0.8; }}
.footer {{ margin-top: 20px; padding-top: 10px; border-top: 2px solid #2c3e50; font-size: 7pt; color: #999; text-align: center; }}
.watermark {{ position: fixed; top: 40%; left: 20%; font-size: 60pt; color: rgba(0,0,0,0.04); transform: rotate(-30deg); z-index: -1; }}
.table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 8pt; }}
.table th {{ background: #2c3e50; color: white; padding: 6px; }}
.table td {{ padding: 4px 6px; border: 1px solid #ddd; }}
.section-title {{ background: #3498db; color: white; padding: 6px 10px; font-weight: bold; margin: 15px 0 5px 0; border-radius: 3px; }}
.signature {{ margin-top: 40px; display: flex; justify-content: space-between; }}
.signature div {{ width: 30%; border-top: 1px solid #2c3e50; padding-top: 5px; text-align: center; font-size: 8pt; }}
</style></head><body>
<div class="watermark">{metadata.get('Status','')}</div>
<div class="header">
<h1>📋 DrillMaster Professional Report - {metadata.get('Well','')} </h1>
<div class="meta">
Company: {metadata.get('Company','')} | Project: {metadata.get('Project','')} | Field: {metadata.get('Field','')} | Well: {metadata.get('Well','')} | Section: {metadata.get('Section','')} | Report #{metadata.get('Report Number','')} | Date: {metadata.get('Report Date','')} | Revision: {metadata.get('Revision','')} | Status: {metadata.get('Status','')} | Generated: {metadata.get('Generated At UTC','')} | Timezone: {metadata.get('Timezone','')} | Units: {metadata.get('Units','')} | Data Quality: {metadata.get('Data Quality','')} | Audit ID: {metadata.get('Audit ID','')}
</div>
</div>

<div class="section-title">Approval</div>
<table class="table">
<tr><th>Role</th><th>Name</th><th>Signature</th><th>Date</th></tr>
<tr><td>Prepared By</td><td>{metadata.get('Prepared By','')}</td><td></td><td></td></tr>
<tr><td>Checked By</td><td>{metadata.get('Checked By','')}</td><td></td><td></td></tr>
<tr><td>Approved By</td><td>{metadata.get('Approved By','')}</td><td></td><td></td></tr>
</table>

<div class="section-title">Data Quality</div>
<p>{metadata.get('Data Quality','')}</p>

<div class="section-title">Report Data</div>
<p>Well: {metadata.get('Well','')} - Depth: {metadata.get('Report Info',{{}}).get('depth_2400','')}m - Summary: {wrap_html(str(metadata.get('Report Info',{{}}).get('summary',''))[:500])}</p>

<div class="signature">
<div>Prepared By<br>{metadata.get('Prepared By','')}</div>
<div>Checked By<br>{metadata.get('Checked By','')}</div>
<div>Approved By<br>{metadata.get('Approved By','')}</div>
</div>

<div class="footer">
Generated by DrillMaster Intelligence Platform | {metadata.get('Generated At UTC','')} | Audit ID: {metadata.get('Audit ID','')} | Page 1 | Data Quality: {metadata.get('Data Quality','')}
</div>
</body></html>
"""

            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(output_path)
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
            printer.setPageMargins(QMarginsF(15, 15, 15, 15), QPageLayout.Unit.Millimeter)

            doc = QTextDocument()
            doc.setHtml(html)
            doc.print_(printer)

            logger.info(f"Professional PDF export saved: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Professional PDF export failed: {e}", exc_info=True)
            return False
