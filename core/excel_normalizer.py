"""Normalize XLSX presentation quirks before semantic import."""
from pathlib import Path
from copy import copy
from openpyxl import load_workbook


class ExcelNormalizationError(Exception):
    pass


def normalize_xlsx(source, destination=None, *, unhide=True, remove_empty=True):
    source = Path(source)
    destination = Path(destination or source.with_name(source.stem + "_clean.xlsx"))
    try:
        workbook = load_workbook(source, data_only=True)
        report = {"source": str(source), "destination": str(destination), "sheets": [], "merged_ranges": 0, "filled_cells": 0}
        for sheet in workbook.worksheets:
            merged = list(sheet.merged_cells.ranges)
            for cell_range in merged:
                min_col, min_row, max_col, max_row = cell_range.bounds
                master = sheet.cell(min_row, min_col)
                value, style = master.value, copy(master._style)
                # MergedCell proxies are read-only. Unmerge first, then write
                # into ordinary Cell objects.
                sheet.unmerge_cells(str(cell_range))
                for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        if cell.value != value:
                            cell.value = value
                            report["filled_cells"] += 1
                        if style:
                            cell._style = copy(style)
                report["merged_ranges"] += 1
            if unhide:
                for dim in sheet.row_dimensions.values():
                    dim.hidden = False
                for dim in sheet.column_dimensions.values():
                    dim.hidden = False
            if remove_empty:
                while sheet.max_row and all(sheet.cell(sheet.max_row, c).value is None for c in range(1, sheet.max_column + 1)):
                    sheet.delete_rows(sheet.max_row, 1)
                while sheet.max_column and all(sheet.cell(r, sheet.max_column).value is None for r in range(1, sheet.max_row + 1)):
                    sheet.delete_cols(sheet.max_column, 1)
            report["sheets"].append({"name": sheet.title, "merged": len(merged), "rows": sheet.max_row, "columns": sheet.max_column})
        workbook.save(destination)
        workbook.close()
        return report
    except Exception as exc:
        raise ExcelNormalizationError(str(exc)) from exc
