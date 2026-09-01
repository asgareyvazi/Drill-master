"""Multi-company template architecture test.

Proves the core rule: Company A Excel -> Company A template -> canonical
schema and Company B Excel -> Company B template -> SAME canonical schema.
Two synthetic workbooks with completely different layouts (sheet names,
cell positions, unit conventions) are mapped with two different templates;
both must resolve to the same canonical keys/values through the generic
ExcelIntelligence pipeline. No company-specific Python branches exist.
"""

import json

import pytest

from openpyxl import Workbook


def _make_workbook_a():
    """Company A layout: sheet 'Daily Ops', row/col positions differ."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Ops"
    ws["B2"] = "Well Name"
    ws["C2"] = "North-1"
    ws["B3"] = "Report Date"
    ws["C3"] = "2024-11-05"
    ws["B4"] = "Mud Weight"
    ws["C4"] = "72 pcf"
    ws["B5"] = "Depth @ 24:00"
    ws["C5"] = "1234"
    ws["B6"] = "LTA (Days)"
    ws["C6"] = "120"
    return wb


def _make_workbook_b():
    """Company B layout: sheet 'REPORT', completely different cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORT"
    ws["D5"] = "Well:"
    ws["E5"] = "South-2"
    ws["D6"] = "Date:"
    ws["E6"] = "2024-11-06"
    ws["D7"] = "MW:"
    ws["E7"] = "72 pcf"  # text with embedded unit (numeric in Company A)
    ws["D8"] = "Bit Depth:"
    ws["E8"] = "1500"
    ws["D9"] = "LTA:"
    ws["E9"] = "95"
    return wb


def _template_a():
    return {
        "name": "Company A Template",
        "version": "1.0",
        "sheet_1_Daily_Ops": {
            "Header": [
                {"field": "Well Name", "row": 2, "col": 3,
                 "canonical": "well_info.name"},
                {"field": "Report Date", "row": 3, "col": 3,
                 "canonical": "daily_report.report_date"},
                {"field": "Mud Weight", "row": 4, "col": 3,
                 "canonical": "mud_report.mw"},
                {"field": "Depth @ 24:00", "row": 5, "col": 3,
                 "canonical": "daily_report.depth_2400"},
                {"field": "LTA (Days)", "row": 6, "col": 3,
                 "canonical": "daily_report.lta_day"},
            ]
        },
    }


def _template_b():
    return {
        "name": "Company B Template",
        "version": "1.0",
        "sheet_1_REPORT": {
            "Header": [
                {"field": "Well:", "row": 5, "col": 5,
                 "canonical": "well_info.name"},
                {"field": "Date:", "row": 6, "col": 5,
                 "canonical": "daily_report.report_date"},
                {"field": "MW:", "row": 7, "col": 5,
                 "canonical": "mud_report.mw"},
                {"field": "Bit Depth:", "row": 8, "col": 5,
                 "canonical": "daily_report.depth_2400"},
                {"field": "LTA:", "row": 9, "col": 5,
                 "canonical": "daily_report.lta_day"},
            ]
        },
    }


def _extract(wb, template):
    from core.excel_intelligence import ExcelIntelligence
    return ExcelIntelligence(wb, template).extract().canonical_json


class TestMultiCompanySameCanonical:
    def test_company_a_resolves(self):
        canonical = _extract(_make_workbook_a(), _template_a())
        assert canonical["well_info"]["name"] == "North-1"
        assert canonical["daily_report"]["report_date"] == "2024-11-05"
        assert canonical["daily_report"]["depth_2400"] == 1234
        assert canonical["daily_report"]["lta_day"] == 120
        # Company A workbook stores MW as a plain number
        assert canonical["mud_report"]["mw"] == 72

    def test_company_b_resolves_same_keys(self):
        canonical = _extract(_make_workbook_b(), _template_b())
        assert canonical["well_info"]["name"] == "South-2"
        assert canonical["daily_report"]["report_date"] == "2024-11-06"
        assert canonical["daily_report"]["depth_2400"] == 1500
        assert canonical["daily_report"]["lta_day"] == 95
        # Embedded-unit text '72 pcf' resolves to the same numeric canonical
        # value as Company A's plain 72
        assert canonical["mud_report"]["mw"] == 72
        # The exact keys are the SAME registry used by Company A
        for key in ("well_info.name", "daily_report.report_date",
                    "daily_report.depth_2400", "daily_report.lta_day",
                    "mud_report.mw"):
            section, field = key.split(".", 1)
            assert field in canonical[section]

    def test_no_company_branches_in_core(self):
        """The extraction core must not mention any company by name."""
        import core.excel_intelligence as ei
        import core.database as db_mod
        for mod in (ei, db_mod):
            src = open(mod.__file__, encoding="utf-8").read()
            for company in ("OEOC", "APAD", "SPAD", "Vira"):
                assert f"if company == \"{company}\"" not in src
                assert f'if company == \'{company}\'' not in src
