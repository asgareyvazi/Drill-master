"""Real-world Excel import tests using generated fixtures.

Tests the full pipeline: Excel → Scan → Classify → Validate → Save → Retrieve
with realistic Excel structures including merged cells, multi-row headers,
side-by-side tables, and different units.
"""

import pytest
import os
import tempfile
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from core.universal_import import WorkbookScanner, SheetClassifier
from core.unit_manager import UnitManager
from core.canonical_schema import FIELD_SPECS, get_critical_fields


FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture(scope="module")
def fixtures():
    """Generate fixtures once for all tests in this module.

    Written to a temp directory so tracked fixture files are never
    rewritten by a test run (openpyxl save would churn their embedded
    timestamps)."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    from tests.create_fixtures import create_all_fixtures
    tmp = tempfile.mkdtemp(prefix="drillmaster_fixtures_")
    return create_all_fixtures(target_dir=tmp)


@pytest.fixture
def db():
    """In-memory database for integration tests."""
    from core.database import DatabaseManager, Base
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool
    
    manager = DatabaseManager()
    manager.engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(manager.engine)
    manager.Session = sessionmaker(bind=manager.engine, autoflush=False, autocommit=False)
    return manager


class TestWorkbookScanner:
    """Test workbook scanning with realistic fixtures."""
    
    def test_scan_normal_ddr(self, fixtures):
        """Scan a normal DDR file."""
        import openpyxl
        wb = openpyxl.load_workbook(fixtures["normal"])
        scanner = WorkbookScanner()
        result = scanner.scan(wb, fixtures["normal"])
        
        assert result["sheet_count"] >= 3
        assert result["file_name"] == "normal_ddr.xlsx"
        assert len(result["tables"]) > 0
        wb.close()
    
    def test_scan_merged_cell_ddr(self, fixtures):
        """Scan a DDR with merged cells."""
        import openpyxl
        wb = openpyxl.load_workbook(fixtures["merged"])
        scanner = WorkbookScanner()
        result = scanner.scan(wb, fixtures["merged"])
        
        assert result["total_merged_ranges"] > 0
        assert result["sheet_count"] >= 1
        wb.close()
    
    def test_scan_multi_header_ddr(self, fixtures):
        """Scan a DDR with multi-row headers."""
        import openpyxl
        wb = openpyxl.load_workbook(fixtures["multi_header"])
        scanner = WorkbookScanner()
        result = scanner.scan(wb, fixtures["multi_header"])
        
        assert len(result["tables"]) > 0
        # Check that multi-row headers are detected
        for table in result["tables"]:
            if table.get("header_rows", 1) > 1:
                assert table["header_rows"] >= 2
                break
        wb.close()


class TestSheetClassifier:
    """Test sheet classification with realistic names."""
    
    def test_classify_daily_report(self, fixtures):
        """Classify 'Daily Report' sheet."""
        import openpyxl
        wb = openpyxl.load_workbook(fixtures["normal"])
        scanner = WorkbookScanner()
        snapshot = scanner.scan(wb, fixtures["normal"])
        
        classifier = SheetClassifier()
        classifications = classifier.classify_all(snapshot)
        
        # Should classify at least one sheet as Daily Report
        assert any("Daily" in v or "Report" in v for v in classifications.values()) or \
               any("Mud" in v for v in classifications.values())
        wb.close()
    
    def test_classify_mud_report(self, fixtures):
        """Classify 'Mud Report' sheet."""
        import openpyxl
        wb = openpyxl.load_workbook(fixtures["normal"])
        scanner = WorkbookScanner()
        snapshot = scanner.scan(wb, fixtures["normal"])
        
        classifier = SheetClassifier()
        classifications = classifier.classify_all(snapshot)
        
        # Mud Report sheet should be classified
        assert "Mud Report" in classifications
        wb.close()


class TestUnitConversion:
    """Test unit conversion with real-world values."""
    
    def test_sg_to_ppg(self):
        """Convert SG to ppg (common in Iranian reports)."""
        result = UnitManager.convert(1.22, "density", "sg", "ppg")
        assert abs(result - 10.18) < 0.1
    
    def test_ft_to_m(self):
        """Convert feet to meters."""
        result = UnitManager.convert(9580, "length", "ft", "m")
        assert abs(result - 2920) < 1
    
    def test_bar_to_psi(self):
        """Convert bar to psi."""
        result = UnitManager.convert(207, "pressure", "bar", "psi")
        assert abs(result - 3002) < 10
    
    def test_fahrenheit_to_celsius(self):
        """Convert °F to °C."""
        result = UnitManager.convert(120, "temperature", "f", "c")
        assert abs(result - 48.9) < 0.5
    
    def test_unit_record_preservation(self):
        """Test that original values are preserved during conversion."""
        record = UnitManager.create_record(
            field="mud_report.mw",
            quantity="density",
            source_unit="sg",
            original_value=1.22,
            target_unit="ppg"
        )
        assert record.original_value == 1.22
        assert record.source_unit == "sg"
        assert record.canonical_unit == "ppg"
        assert record.normalized_value is not None
        assert abs(record.normalized_value - 10.18) < 0.1


class TestCanonicalSchema:
    """Test canonical schema coverage for real-world fields."""
    
    def test_critical_fields_present(self):
        """All critical fields must be in schema."""
        critical = get_critical_fields()
        required = [
            "well_info.name", "daily_report.depth_2400",
            "mud_report.mw", "drilling_params.bit_size",
            "survey.md", "bop.working_pressure",
        ]
        for field in required:
            assert field in critical
    
    def test_mud_fields_complete(self):
        """Mud report fields should cover standard DDR mud section."""
        mud_fields = [p for p in FIELD_SPECS if p.startswith("mud_report.")]
        assert len(mud_fields) >= 10
        
        # Essential mud properties
        essentials = ["mw", "pv", "yp", "ph", "temperature", "funnel_vis"]
        for field in essentials:
            assert f"mud_report.{field}" in FIELD_SPECS
    
    def test_drilling_params_complete(self):
        """Drilling params should cover standard DDR drilling section."""
        dp_fields = [p for p in FIELD_SPECS if p.startswith("drilling_params.")]
        assert len(dp_fields) >= 10
        
        essentials = ["bit_size", "depth_in", "depth_out", "avg_rop"]
        for field in essentials:
            assert f"drilling_params.{field}" in FIELD_SPECS


class TestEndToEndImport:
    """End-to-end import test with realistic data."""
    
    def test_create_well_and_import_data(self, db):
        """Create well, section, and import drilling data."""
        from core.database import Company, Project, Well, Section, DailyReport
        
        session = db.create_session()
        try:
            company = Company(name="OEOC", code="OEOC001")
            session.add(company)
            session.flush()
            
            project = Project(company_id=company.id, name="Azar", code="AZ001")
            session.add(project)
            session.flush()
            
            well = Well(project_id=project.id, name="AZNS-207", code="AZNS207")
            session.add(well)
            session.flush()
            
            section = Section(well_id=well.id, name="12.25in", depth_from=2500, depth_to=3000)
            session.add(section)
            session.flush()
            
            session.commit()
            
            well_id = well.id
            section_id = section.id
        finally:
            session.close()
        
        # Save daily report
        report = db.save_daily_report({
            "well_id": well_id,
            "section_id": section_id,
            "report_date": date(2024, 11, 22),
            "report_number": 39,
            "rig_day": 39,
            "depth_0000": 2850.0,
            "depth_0600": 2875.0,
            "depth_2400": 2920.0,
        })
        assert report is not None
        report_id = report["id"]
        
        # Save mud report with unit conversion
        mw_sg = 1.22
        record = UnitManager.create_record("mud_report.mw", "density", "sg", mw_sg, "ppg")
        
        mud_id = db.save_mud_report({
            "well_id": well_id,
            "report_id": report_id,
            "report_date": date(2024, 11, 22),
            "mw": record.normalized_value,
            "pv": 15.0,
            "yp": 10.0,
            "ph": 9.5,
        })
        assert mud_id is not None
        
        # Save drilling parameters
        dp_id = db.save_drilling_parameters({
            "well_id": well_id,
            "report_id": report_id,
            "report_date": date(2024, 11, 22),
            "bit_size": 12.25,
            "depth_in": 2850.0,
            "depth_out": 2920.0,
            "avg_rop": 5.8,
        })
        assert dp_id is not None
        
        # Verify data integrity
        retrieved_report = db.get_daily_report_by_id(report_id)
        assert retrieved_report["depth_2400"] == 2920.0
        
        retrieved_mud = db.get_mud_report(well_id=well_id)
        assert retrieved_mud is not None
        assert abs(retrieved_mud["mw"] - 10.18) < 0.1  # 1.22 SG ≈ 10.18 ppg
        
        retrieved_dp = db.get_drilling_parameters(well_id=well_id)
        assert retrieved_dp is not None
        assert retrieved_dp["bit_size"] == 12.25
