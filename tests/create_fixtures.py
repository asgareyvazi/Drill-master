"""Generate realistic Excel test fixtures for integration testing.

Creates Excel files that mimic real-world drilling reports with:
- Merged cells
- Multi-row headers
- Multiple tables per sheet
- Side-by-side tables
- Hidden rows/columns
- Formula cells
- Different units
- Inconsistent naming
"""

import os
from pathlib import Path
from datetime import date, time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


FIXTURES_DIR = Path(__file__).parent / "fixtures"


def create_normal_ddr(filepath: str = None) -> str:
    """Create a normal Daily Drilling Report Excel file."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for fixture generation")
    
    filepath = filepath or str(FIXTURES_DIR / "normal_ddr.xlsx")
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    
    # Sheet 1: Daily Report
    ws = wb.active
    ws.title = "Daily Report"
    
    # Header section
    ws["A1"] = "DAILY DRILLING REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Well Name:"
    ws["B3"] = "AZNS-207"
    ws["A4"] = "Report Date:"
    ws["B4"] = date(2024, 11, 22)
    ws["A5"] = "Report Number:"
    ws["B5"] = 39
    ws["A6"] = "Rig Day:"
    ws["B6"] = 39
    
    # Depth section
    ws["A8"] = "Depth @ 00:00"
    ws["B8"] = 2850.0
    ws["A9"] = "Depth @ 06:00"
    ws["B9"] = 2875.0
    ws["A10"] = "Depth @ 24:00"
    ws["B10"] = 2920.0
    
    # Time Log section
    ws["D1"] = "TIME LOG (24H)"
    ws["D1"].font = Font(bold=True)
    ws["D2"] = "From"
    ws["E2"] = "To"
    ws["F2"] = "Duration"
    ws["G2"] = "Main Code"
    ws["H2"] = "Sub Code"
    ws["I2"] = "Description"
    ws["J2"] = "Contractor"
    
    time_logs = [
        (time(0, 0), time(6, 0), 6.0, "Drilling", "DR-01", "Drilling 12.25in hole", "Nabors"),
        (time(6, 0), time(7, 0), 1.0, "Connection", "CN-01", "Making connection", "Nabors"),
        (time(7, 0), time(12, 0), 5.0, "Drilling", "DR-01", "Drilling 12.25in hole", "Nabors"),
        (time(12, 0), time(13, 0), 1.0, "Trip", "TR-01", "Trip out for bit change", "Nabors"),
        (time(13, 0), time(18, 0), 5.0, "Trip", "TR-01", "Running new BHA", "Nabors"),
        (time(18, 0), "24:00", 6.0, "Drilling", "DR-02", "Drilling with new bit", "Nabors"),
    ]
    
    for i, (tf, tt, dur, mc, sc, desc, cont) in enumerate(time_logs, 3):
        ws[f"D{i}"] = tf
        ws[f"E{i}"] = tt
        ws[f"F{i}"] = dur
        ws[f"G{i}"] = mc
        ws[f"H{i}"] = sc
        ws[f"I{i}"] = desc
        ws[f"J{i}"] = cont
    
    # Sheet 2: Mud Report
    ws2 = wb.create_sheet("Mud Report")
    ws2["A1"] = "MUD REPORT"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Mud Weight (ppg):"
    ws2["B3"] = 10.2
    ws2["A4"] = "PV (cp):"
    ws2["B4"] = 15.0
    ws2["A5"] = "YP (lb/100ft²):"
    ws2["B5"] = 10.0
    ws2["A6"] = "Funnel Viscosity (sec):"
    ws2["B6"] = 38.0
    ws2["A7"] = "Gel 10s:"
    ws2["B7"] = 4.0
    ws2["A8"] = "Gel 10m:"
    ws2["B8"] = 8.0
    ws2["A9"] = "pH:"
    ws2["B9"] = 9.5
    ws2["A10"] = "Temperature (°F):"
    ws2["B10"] = 120.0
    
    # Sheet 3: Survey
    ws3 = wb.create_sheet("Survey")
    ws3["A1"] = "SURVEY DATA"
    ws3["A1"].font = Font(bold=True)
    ws3["A2"] = "MD (m)"
    ws3["B2"] = "Inc (deg)"
    ws3["C2"] = "Azi (deg)"
    ws3["D2"] = "TVD (m)"
    
    surveys = [
        (2800, 2.5, 45.0, 2799.5),
        (2825, 3.0, 46.0, 2824.3),
        (2850, 3.5, 47.0, 2849.0),
        (2875, 4.0, 48.0, 2873.5),
        (2900, 4.5, 49.0, 2897.8),
        (2920, 5.0, 50.0, 2917.0),
    ]
    
    for i, (md, inc, azi, tvd) in enumerate(surveys, 3):
        ws3[f"A{i}"] = md
        ws3[f"B{i}"] = inc
        ws3[f"C{i}"] = azi
        ws3[f"D{i}"] = tvd
    
    wb.save(filepath)
    return filepath


def create_merged_cell_ddr(filepath: str = None) -> str:
    """Create a DDR with merged cells (common in real reports)."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    filepath = filepath or str(FIXTURES_DIR / "merged_cell_ddr.xlsx")
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    
    # Merged title
    ws.merge_cells("A1:C1")
    ws["A1"] = "DAILY DRILLING REPORT - OEOC"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Merged header row
    ws.merge_cells("A3:B3")
    ws["A3"] = "Well Information"
    ws["A3"].font = Font(bold=True)
    
    ws["A4"] = "Well Name"
    ws["B4"] = "OEOC-208"
    ws["A5"] = "Date"
    ws["B5"] = date(2024, 11, 22)
    
    # Merged section header - set value BEFORE merging
    ws["D1"] = "TIME LOG"
    ws.merge_cells("D1:J1")
    ws["D1"].font = Font(bold=True, size=12)
    
    # Multi-row header (merged)
    ws.merge_cells("D2:E2")
    ws["D2"] = "Time Period"
    ws["F2"] = "Hours"
    ws["G2"] = "Activity"
    
    ws["D3"] = "From"
    ws["E3"] = "To"
    
    # Data
    ws["D4"] = time(0, 0)
    ws["E4"] = time(6, 0)
    ws["F4"] = 6.0
    ws["G4"] = "Drilling"
    
    ws["D5"] = time(6, 0)
    ws["E5"] = time(12, 0)
    ws["F5"] = 6.0
    ws["G5"] = "Drilling"
    
    ws["D6"] = time(12, 0)
    ws["E6"] = time(18, 0)
    ws["F6"] = 6.0
    ws["G6"] = "Trip"
    
    ws["D7"] = time(18, 0)
    ws["E7"] = "24:00"
    ws["F7"] = 6.0
    ws["G7"] = "Drilling"
    
    # Mud section (separate area, not overlapping with TIME LOG merge)
    ws["L1"] = "MUD PROPERTIES"
    ws["L1"].font = Font(bold=True)
    ws["L2"] = "Parameter"
    ws["M2"] = "Value"
    ws["L3"] = "MW (ppg)"
    ws["M3"] = 10.5
    ws["L4"] = "PV (cp)"
    ws["M4"] = 18.0
    ws["L5"] = "YP (lbf/100ft²)"
    ws["M5"] = 12.0
    
    wb.save(filepath)
    return filepath


def create_multi_header_ddr(filepath: str = None) -> str:
    """Create a DDR with 2-3 row headers (common in Iranian/NOC reports)."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    filepath = filepath or str(FIXTURES_DIR / "multi_header_ddr.xlsx")
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Drilling Parameters"
    
    # 3-row header
    ws["A1"] = "Bit"
    ws["A2"] = "Number"
    ws["A3"] = "-"
    ws["B1"] = "Bit"
    ws["B2"] = "Size"
    ws["B3"] = "inch"
    ws["C1"] = "Bit"
    ws["C2"] = "Type"
    ws["C3"] = "-"
    ws["D1"] = "Depth"
    ws["D2"] = "In"
    ws["D3"] = "m"
    ws["E1"] = "Depth"
    ws["E2"] = "Out"
    ws["E3"] = "m"
    ws["F1"] = "Hours"
    ws["F2"] = "On Bottom"
    ws["F3"] = "hr"
    ws["G1"] = "ROP"
    ws["G2"] = "Average"
    ws["G3"] = "m/hr"
    
    # Data
    ws["A4"] = "BD-1"
    ws["B4"] = 12.25
    ws["C4"] = "PDC"
    ws["D4"] = 2500
    ws["E4"] = 2850
    ws["F4"] = 48.0
    ws["G4"] = 7.3
    
    ws["A5"] = "BD-2"
    ws["B5"] = 12.25
    ws["C5"] = "PDC"
    ws["D5"] = 2850
    ws["E5"] = 2920
    ws["F5"] = 12.0
    ws["G5"] = 5.8
    
    wb.save(filepath)
    return filepath


def create_unit_variant_ddr(filepath: str = None) -> str:
    """Create a DDR using non-standard units (SG instead of ppg, ft instead of m)."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    filepath = filepath or str(FIXTURES_DIR / "unit_variant_ddr.xlsx")
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    
    ws["A1"] = "Well:"
    ws["B1"] = "Test Well"
    ws["A2"] = "Date:"
    ws["B2"] = date(2024, 11, 22)
    
    # Using SG instead of ppg
    ws["A4"] = "Mud Weight (SG):"
    ws["B4"] = 1.22  # ~10.2 ppg
    
    # Using ft instead of m
    ws["A5"] = "Depth (ft):"
    ws["B5"] = 9580.0  # ~2920 m
    
    # Using bar instead of psi
    ws["A6"] = "Pump Pressure (bar):"
    ws["B6"] = 207.0  # ~3000 psi
    
    # Using °C
    ws["A7"] = "Temperature (°C):"
    ws["B7"] = 49.0  # ~120°F
    
    wb.save(filepath)
    return filepath


def create_all_fixtures(target_dir: str = None):
    """Generate all test fixtures.

    Writes into FIXTURES_DIR when no target_dir is given (CLI use), or
    into target_dir when provided — so test runs never rewrite the
    tracked fixture files (openpyxl save would otherwise churn
    docProps/core.xml timestamps on every run).
    """
    fixtures = {}
    fixtures["normal"] = create_normal_ddr(
        str(Path(target_dir) / "normal_ddr.xlsx") if target_dir else None
    )
    fixtures["merged"] = create_merged_cell_ddr(
        str(Path(target_dir) / "merged_cell_ddr.xlsx") if target_dir else None
    )
    fixtures["multi_header"] = create_multi_header_ddr(
        str(Path(target_dir) / "multi_header_ddr.xlsx") if target_dir else None
    )
    fixtures["unit_variant"] = create_unit_variant_ddr(
        str(Path(target_dir) / "unit_variant_ddr.xlsx") if target_dir else None
    )
    return fixtures


if __name__ == "__main__":
    paths = create_all_fixtures()
    for name, path in paths.items():
        print(f"Created: {name} -> {path}")
