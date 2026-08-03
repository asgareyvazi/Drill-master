# dialogs/planning_dialog.py

"""
Well Plan Dialog - برنامه‌ریزی حرفه‌ای عملیات حفاری
شامل: Phase/Category management, Activity timeline, PLAN vs FACT
"""
import json
import logging
from datetime import datetime, date, timedelta

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

from core.database import DatabaseManager, Well, Section, PlannedActivity, WellPlan

logger = logging.getLogger(__name__)
# ==================== Plan Import Review Dialog ====================
class PlanImportReviewDialog(QDialog):
    """دیالوگ Review برای Import Plan از Excel"""

    PLAN_FIELDS = [
        ("activity", "Activity Description", True),
        ("iadc_code", "IADC Code", True),
        ("interval", "Interval (m)", False),
        ("formation", "Formation", False),
        ("depth", "Depth (m)", True),
        ("rop", "ROP", False),
        ("hours", "Duration (Hours)", True),
        ("days", "Duration (Days)", False),
        ("total_days", "Total Duration (Days)", False),
    ]

    def __init__(self, filepath, parent=None):
        super().__init__(parent)
        self.filepath = filepath
        self.col_map = {}
        self.cell_cache = {}
        self.activities = []
        self._selected_field = None
        self._pending_col = None

        self.setWindowTitle("📋 Plan Import - Review & Assign Columns")
        self.setMinimumSize(1100, 650)
        self.setModal(True)

        self._load_excel()
        self._init_ui()
        self._auto_detect_columns()

    def _load_excel(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb.active
        self.sheet_name = ws.title
        self.cell_cache = {}
        self.max_row = min(ws.max_row or 1, 200)
        self.max_col = min(ws.max_column or 1, 20)

        for r in range(1, self.max_row + 1):
            for c in range(1, self.max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    self.cell_cache[(r, c)] = val

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(3)

        # Header
        header = QLabel(
            f"📋 Assign columns from: {self.filepath.split('/')[-1].split(chr(92))[-1]}"
        )
        header.setStyleSheet(
            "font-size: 11px; font-weight: bold; padding: 4px; "
            "background: #d5f5e3; border-radius: 3px; color: #27ae60;"
        )
        header.setFixedHeight(26)
        layout.addWidget(header)

        legend = QLabel(
            "🟢 Auto-detected | 🔴 Not found → click column header in Excel"
        )
        legend.setStyleSheet("color: #888; font-size: 9px;")
        legend.setFixedHeight(14)
        layout.addWidget(legend)

        splitter = QSplitter(Qt.Horizontal)

        # ===== LEFT: Field Assignment =====
        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.setSpacing(2)

        ll.addWidget(QLabel("<b>Column Assignments:</b>"))

        self.field_table = QTableWidget(0, 4)
        self.field_table.setHorizontalHeaderLabels([
            "", "Field", "Column", "Sample"
        ])
        self.field_table.setColumnWidth(0, 28)
        self.field_table.setColumnWidth(1, 150)
        self.field_table.setColumnWidth(2, 60)
        self.field_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.Stretch
        )
        self.field_table.verticalHeader().setDefaultSectionSize(22)
        self.field_table.verticalHeader().setVisible(False)
        self.field_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.field_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.field_table.itemClicked.connect(self._on_field_clicked)
        self.field_table.setStyleSheet(
            "QTableWidget{font-size:10px;}"
            "QHeaderView::section{background:#34495e;color:white;"
            "font-size:9px;padding:3px;}"
        )
        ll.addWidget(self.field_table)

        # Header row selector
        hr_layout = QHBoxLayout()
        hr_layout.addWidget(QLabel("Header Row:"))
        self.header_row_spin = QSpinBox()
        self.header_row_spin.setRange(1, 20)
        self.header_row_spin.setValue(1)
        self.header_row_spin.valueChanged.connect(self._auto_detect_columns)
        hr_layout.addWidget(self.header_row_spin)

        hr_layout.addWidget(QLabel("Data Start:"))
        self.data_start_spin = QSpinBox()
        self.data_start_spin.setRange(2, 50)
        self.data_start_spin.setValue(2)
        hr_layout.addWidget(self.data_start_spin)
        hr_layout.addStretch()
        ll.addLayout(hr_layout)

        # Assign button
        self.assign_btn = QPushButton("✅ Assign Selected Column")
        self.assign_btn.setStyleSheet(
            "background:#27ae60;color:white;font-weight:bold;"
            "padding:6px;border-radius:3px;border:none;"
        )
        self.assign_btn.setEnabled(False)
        self.assign_btn.clicked.connect(self._assign_column)
        ll.addWidget(self.assign_btn)

        splitter.addWidget(left)

        # ===== RIGHT: Excel Preview =====
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(2)

        rl.addWidget(QLabel(f"<b>📄 {self.sheet_name}</b>"))

        self.excel_table = QTableWidget()
        self.excel_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.excel_table.setSelectionMode(QTableWidget.SingleSelection)
        self.excel_table.setSelectionBehavior(QTableWidget.SelectColumns)
        self.excel_table.verticalHeader().setDefaultSectionSize(18)
        self.excel_table.setStyleSheet(
            "QTableWidget{font-size:9px;}"
            "QTableWidget::item:selected{background:#e74c3c;color:white;}"
            "QHeaderView::section{background:#2c3e50;color:white;"
            "font-size:8px;padding:2px;}"
        )
        self.excel_table.cellClicked.connect(self._on_excel_clicked)
        rl.addWidget(self.excel_table)

        self.cell_info = QLabel("Click a column to assign")
        self.cell_info.setStyleSheet(
            "color:#666;font-size:9px;padding:3px;"
            "background:#f0f0f0;border-radius:2px;"
        )
        self.cell_info.setFixedHeight(28)
        rl.addWidget(self.cell_info)

        splitter.addWidget(right)
        splitter.setSizes([350, 650])
        layout.addWidget(splitter)

        # Bottom
        bl = QHBoxLayout()
        bl.setSpacing(5)

        auto_btn = QPushButton("🤖 Re-Detect")
        auto_btn.clicked.connect(self._auto_detect_columns)
        bl.addWidget(auto_btn)

        clear_btn = QPushButton("🗑️ Clear All")
        clear_btn.clicked.connect(self._clear_assignments)
        bl.addWidget(clear_btn)

        bl.addStretch()

        preview_btn = QPushButton("👁️ Preview Data")
        preview_btn.setStyleSheet(
            "background:#3498db;color:white;font-weight:bold;"
            "padding:6px 12px;border-radius:3px;border:none;"
        )
        preview_btn.clicked.connect(self._preview_data)
        bl.addWidget(preview_btn)

        import_btn = QPushButton("🚀 Import Plan")
        import_btn.setStyleSheet(
            "background:#27ae60;color:white;font-weight:bold;"
            "font-size:11px;padding:6px 16px;border-radius:3px;border:none;"
        )
        import_btn.clicked.connect(self._do_import)
        bl.addWidget(import_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        bl.addWidget(cancel_btn)

        layout.addLayout(bl)

        self._display_excel()
        self._populate_field_table()

    def _display_excel(self):
        self.excel_table.setRowCount(min(self.max_row, 100))
        self.excel_table.setColumnCount(self.max_col)

        hdrs = [
            chr(64 + c) if c <= 26
            else f"{chr(64 + (c-1)//26)}{chr(65 + (c-1)%26)}"
            for c in range(1, self.max_col + 1)
        ]
        self.excel_table.setHorizontalHeaderLabels(hdrs)

        for (r, c), val in self.cell_cache.items():
            if r > 100 or c > self.max_col:
                continue
            item = QTableWidgetItem(str(val)[:60])
            item.setToolTip(str(val))
            self.excel_table.setItem(r - 1, c - 1, item)

        self.excel_table.resizeColumnsToContents()
        for c in range(self.excel_table.columnCount()):
            w = self.excel_table.columnWidth(c)
            self.excel_table.setColumnWidth(c, max(50, min(w, 200)))

    def _populate_field_table(self):
        self.field_table.setRowCount(0)
        for key, label, required in self.PLAN_FIELDS:
            row = self.field_table.rowCount()
            self.field_table.insertRow(row)

            assigned_col = self.col_map.get(key)
            if assigned_col:
                icon = "🟢"
                bg = QColor("#eafaf1")
                col_letter = chr(64 + assigned_col) if assigned_col <= 26 else "?"
                sample = str(self.cell_cache.get(
                    (self.data_start_spin.value(), assigned_col), ""
                ))[:30]
            else:
                icon = "🔴" if required else "⚪"
                bg = QColor("#fdedec") if required else QColor("#f8f8f8")
                col_letter = "—"
                sample = ""

            si = QTableWidgetItem(icon)
            si.setTextAlignment(Qt.AlignCenter)
            si.setBackground(bg)
            self.field_table.setItem(row, 0, si)

            fi = QTableWidgetItem(f"{'*' if required else ''}{label}")
            fi.setData(Qt.UserRole, key)
            fi.setBackground(bg)
            if required:
                font = fi.font()
                font.setBold(True)
                fi.setFont(font)
            self.field_table.setItem(row, 1, fi)

            ci = QTableWidgetItem(col_letter)
            ci.setTextAlignment(Qt.AlignCenter)
            ci.setBackground(bg)
            self.field_table.setItem(row, 2, ci)

            smi = QTableWidgetItem(sample)
            smi.setForeground(QColor("#666"))
            smi.setBackground(bg)
            self.field_table.setItem(row, 3, smi)

    def _auto_detect_columns(self):
        self.col_map = {}
        hr = self.header_row_spin.value()

        keywords = {
            "activity": ["activity", "description", "planned", "operation"],
            "iadc_code": ["iadc", "code"],
            "interval": ["interval"],
            "formation": ["formation"],
            "depth": ["depth"],
            "rop": ["rop"],
            "hours": ["hour", "time duration"],
            "days": ["day", "duration"],
            "total_days": ["total", "cumulative"],
        }

        for c in range(1, self.max_col + 1):
            val = self.cell_cache.get((hr, c))
            if not val:
                continue
            txt = str(val).lower().strip()

            for field_key, kws in keywords.items():
                if field_key in self.col_map:
                    continue
                for kw in kws:
                    if kw in txt:
                        # خاص: "total" فقط با "day" یا "duration"
                        if field_key == "total_days" and "day" not in txt and "duration" not in txt:
                            continue
                        # خاص: "days" نباید "total" باشد
                        if field_key == "days" and "total" in txt:
                            continue
                        # خاص: hours باید "hour" داشته باشد
                        if field_key == "hours" and "hour" not in txt:
                            continue
                        self.col_map[field_key] = c
                        break

        # Activity: اگر پیدا نشد، اولین ستون با متن بلند
        if "activity" not in self.col_map:
            for c in range(1, self.max_col + 1):
                if c in self.col_map.values():
                    continue
                # چک کنیم آیا ستون متن بلند دارد
                sample = self.cell_cache.get(
                    (self.data_start_spin.value(), c), ""
                )
                if sample and len(str(sample)) > 20:
                    self.col_map["activity"] = c
                    break

        # Data start row auto-detect
        if self.col_map:
            first_col = list(self.col_map.values())[0]
            for r in range(hr + 1, hr + 10):
                val = self.cell_cache.get((r, first_col))
                if val and str(val).strip():
                    self.data_start_spin.setValue(r)
                    break

        self._populate_field_table()
        self._highlight_assigned_columns()

    def _highlight_assigned_columns(self):
        """Highlight ستون‌های assign شده در Excel"""
        assigned_cols = set(self.col_map.values())
        for r in range(self.excel_table.rowCount()):
            for c in range(self.excel_table.columnCount()):
                item = self.excel_table.item(r, c)
                if item:
                    if (c + 1) in assigned_cols:
                        item.setBackground(QColor("#d5f5e3"))
                    else:
                        item.setBackground(QColor("white"))

    def _on_field_clicked(self, item):
        row = item.row()
        fi = self.field_table.item(row, 1)
        if fi:
            self._selected_field = fi.data(Qt.UserRole)
            self.cell_info.setText(
                f"📌 Selected: {fi.text()} → Click a column in Excel"
            )
            self.assign_btn.setEnabled(True)

    def _on_excel_clicked(self, row, col):
        self._pending_col = col + 1
        col_letter = chr(64 + self._pending_col) if self._pending_col <= 26 else "?"

        # نمونه مقادیر
        samples = []
        ds = self.data_start_spin.value()
        for r in range(ds, min(ds + 3, self.max_row + 1)):
            v = self.cell_cache.get((r, self._pending_col))
            if v:
                samples.append(str(v)[:30])

        field_text = ""
        if self._selected_field:
            field_text = f" → {self._selected_field}"

        self.cell_info.setText(
            f"📍 Column {col_letter}{field_text} | "
            f"Samples: {', '.join(samples[:3])}"
        )

    def _assign_column(self):
        if not self._selected_field or not self._pending_col:
            return

        self.col_map[self._selected_field] = self._pending_col
        self._populate_field_table()
        self._highlight_assigned_columns()

        col_letter = chr(64 + self._pending_col) if self._pending_col <= 26 else "?"
        self.cell_info.setText(
            f"✅ {self._selected_field} → Column {col_letter}"
        )

    def _clear_assignments(self):
        self.col_map = {}
        self._populate_field_table()
        self._highlight_assigned_columns()

    def _preview_data(self):
        """پیش‌نمایش داده‌های parse شده"""
        data = self._parse_data()
        if not data:
            QMessageBox.warning(self, "No Data", "No activities found.")
            return

        preview = f"📋 Preview: {len(data)} activities\n\n"
        for i, act in enumerate(data[:10]):
            preview += (
                f"{i+1}. [{act.get('iadc','')}] "
                f"{act.get('activity','')[:60]}\n"
                f"   Depth: {act.get('depth',0)}m | "
                f"Formation: {act.get('formation','')} | "
                f"Duration: {act.get('hours',0):.1f}h\n\n"
            )
        if len(data) > 10:
            preview += f"... and {len(data)-10} more activities"

        QMessageBox.information(self, "Preview", preview)

    def _parse_data(self):
        """Parse داده‌ها بر اساس col_map"""
        if "activity" not in self.col_map:
            return []

        data = []
        ds = self.data_start_spin.value()

        for r in range(ds, self.max_row + 1):
            act_val = self.cell_cache.get(
                (r, self.col_map["activity"])
            )
            if not act_val:
                continue

            activity = str(act_val).strip()
            if len(activity) < 5:
                continue

            # Skip summary rows
            al = activity.lower()
            if any(s in al for s in [
                "total duration", "note:", "contingency plan",
                "based on operation"
            ]):
                continue

            row_data = {"activity": activity}

            for key in ["iadc_code", "interval", "formation",
                        "depth", "rop", "hours", "days", "total_days"]:
                col = self.col_map.get(key)
                if not col:
                    continue
                val = self.cell_cache.get((r, col))
                if val is None:
                    row_data[key] = "" if key in [
                        "iadc_code", "formation"
                    ] else 0
                    continue

                if key in ["iadc_code", "formation"]:
                    row_data[key] = str(val).strip()
                else:
                    try:
                        row_data[key] = float(val)
                    except (ValueError, TypeError):
                        row_data[key] = 0

            data.append(row_data)

        return data

    def _do_import(self):
        """Import و بستن دیالوگ"""
        data = self._parse_data()
        if not data:
            QMessageBox.warning(self, "No Data", "No activities found.")
            return

        self.activities = data
        self.accept()

    def get_activities(self):
        return self.activities
        
# فازها و کتگوری‌های استاندارد حفاری
DRILLING_PHASES = {
    "01-RIG MOVE & RIG UP": {
        "color": "#95a5a6",
        "activities": [
            "Rig Move & Positioning",
            "Rig Up",
            "BOP & Wellhead Installation",
            "Pre-Spud Activities",
        ]
    },
    "02-CONDUCTOR": {
        "color": "#e74c3c",
        "activities": [
            "Drill Conductor Hole",
            "Run Conductor Casing",
            "Cement Conductor",
            "WOC Conductor",
        ]
    },
    "03-SURFACE HOLE": {
        "color": "#e67e22",
        "activities": [
            "Drill Surface Hole",
            "Circulate & Condition",
            "Run Surface Casing",
            "Cement Surface Casing",
            "WOC Surface Casing",
            "Nipple Up BOP",
            "BOP Test",
            "FIT/LOT",
        ]
    },
    "04-INTERMEDIATE HOLE": {
        "color": "#f39c12",
        "activities": [
            "Drill Intermediate Hole",
            "Wiper Trip",
            "Circulate & Condition",
            "Logging",
            "Run Intermediate Casing",
            "Cement Intermediate Casing",
            "WOC Intermediate Casing",
            "BOP Test",
            "FIT/LOT",
        ]
    },
    "05-PRODUCTION HOLE": {
        "color": "#27ae60",
        "activities": [
            "Drill Production Hole",
            "Short Trip",
            "Wiper Trip",
            "Circulate & Condition",
            "Logging (Wireline/LWD)",
            "Run Production Casing",
            "Cement Production Casing",
            "WOC Production Casing",
        ]
    },
    "06-LINER SECTION": {
        "color": "#3498db",
        "activities": [
            "Drill Liner Section",
            "Circulate & Condition",
            "Run Liner",
            "Set Liner Hanger",
            "Cement Liner",
            "WOC Liner",
            "Drill Out Cement",
        ]
    },
    "07-COMPLETION": {
        "color": "#9b59b6",
        "activities": [
            "Run Completion String",
            "Perforation",
            "Acidizing / Stimulation",
            "Flow Test",
            "Install Wellhead / Xmas Tree",
        ]
    },
    "08-TESTING": {
        "color": "#1abc9c",
        "activities": [
            "DST",
            "Production Test",
            "Pressure Build-Up",
            "Clean Up",
        ]
    },
    "09-RIG DOWN & MOVE": {
        "color": "#7f8c8d",
        "activities": [
            "Rig Down",
            "Demobilization",
            "Site Restoration",
        ]
    },
    "10-CONTINGENCY": {
        "color": "#c0392b",
        "activities": [
            "Fishing",
            "Sidetrack",
            "Lost Circulation Treatment",
            "Stuck Pipe Operations",
            "Well Control",
            "Weather Downtime",
        ]
    },
}


class WellPlanDialog(QDialog):
    """دیالوگ حرفه‌ای برنامه حفاری"""

    def __init__(self, db_manager: DatabaseManager, well_id: int, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.well_id = well_id
        self.plan_id = None
        self.activities = []

        self.setWindowTitle(f"📋 Well Drilling Plan")
        self.setMinimumSize(1000, 700)
        self.init_ui()
        self.load_well_info()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(5)

        # Header
        header = QWidget()
        header.setStyleSheet("background: #2c3e50; border-radius: 4px;")
        header.setMaximumHeight(40)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(10, 3, 10, 3)
        self.header_label = QLabel("📋 Well Drilling Plan")
        self.header_label.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        hl.addWidget(self.header_label)
        hl.addStretch()

        save_btn = QPushButton("💾 Save Plan")
        save_btn.setStyleSheet("background: #27ae60; color: white; font-weight: bold; padding: 5px 15px; border-radius: 3px; border: none;")
        save_btn.clicked.connect(self.save_plan)
        hl.addWidget(save_btn)

        cancel_btn = QPushButton("❌ Cancel")
        cancel_btn.setStyleSheet("background: #e74c3c; color: white; padding: 5px 10px; border-radius: 3px; border: none;")
        cancel_btn.clicked.connect(self.reject)
        hl.addWidget(cancel_btn)
        layout.addWidget(header)

        # Tabs
        self.tab_widget = QTabWidget()
        self.tab_widget.addTab(self._create_general_tab(), "📋 General Info")
        self.tab_widget.addTab(self._create_activities_tab(), "📅 Activities")
        self.tab_widget.addTab(self._create_timeline_tab(), "📊 Timeline")
        layout.addWidget(self.tab_widget)

    def _create_general_tab(self):
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QFormLayout(content)
        layout.setSpacing(10)

        self.plan_name = QLineEdit()
        self.plan_name.setPlaceholderText("e.g., Base Plan Rev.0, Revised Plan v2")
        layout.addRow("Plan Name:", self.plan_name)

        self.plan_version = QLineEdit("1.0")
        layout.addRow("Version:", self.plan_version)

        self.spud_date = QDateEdit(QDate.currentDate())
        self.spud_date.setCalendarPopup(True)
        layout.addRow("Planned Spud:", self.spud_date)

        self.finish_date = QDateEdit(QDate.currentDate().addDays(90))
        self.finish_date.setCalendarPopup(True)
        layout.addRow("Planned Finish:", self.finish_date)

        self.final_depth = QDoubleSpinBox()
        self.final_depth.setRange(0, 20000)
        self.final_depth.setSuffix(" m")
        layout.addRow("Final Depth:", self.final_depth)

        self.total_days = QLabel("-- days")
        self.total_days.setStyleSheet("font-weight: bold; color: #3498db; font-size: 13px;")
        layout.addRow("Total Planned Days:", self.total_days)

        self.description = QTextEdit()
        self.description.setMaximumHeight(100)
        self.description.setPlaceholderText("Plan description, assumptions, risks...")
        layout.addRow("Description:", self.description)

        self.spud_date.dateChanged.connect(self._update_total_days)
        self.finish_date.dateChanged.connect(self._update_total_days)
        self._update_total_days()

        scroll.setWidget(content)
        tab_layout = QVBoxLayout(tab)
        tab_layout.addWidget(scroll)
        return tab

    def _update_total_days(self):
        days = self.spud_date.date().daysTo(self.finish_date.date())
        self.total_days.setText(f"{days} days")

    def _create_activities_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # Toolbar
        toolbar = QHBoxLayout()

        add_btn = QPushButton("➕ Add Activity")
        add_btn.setStyleSheet("background: #27ae60; color: white; padding: 5px 12px; border-radius: 3px; border: none; font-weight: bold;")
        add_btn.clicked.connect(self._add_activity_dialog)

        add_phase_btn = QPushButton("📋 Add Phase (Template)")
        add_phase_btn.setStyleSheet("background: #3498db; color: white; padding: 5px 12px; border-radius: 3px; border: none;")
        add_phase_btn.clicked.connect(self._add_phase_template)

        remove_btn = QPushButton("🗑️ Remove")
        remove_btn.clicked.connect(self._remove_activity)

        move_up = QPushButton("⬆️")
        move_up.setFixedWidth(30)
        move_up.clicked.connect(self._move_up)
        move_down = QPushButton("⬇️")
        move_down.setFixedWidth(30)
        move_down.clicked.connect(self._move_down)

        import_btn = QPushButton("📂 Import Excel")
        import_btn.clicked.connect(self._import_from_excel)

        toolbar.addWidget(add_btn)
        toolbar.addWidget(add_phase_btn)
        toolbar.addWidget(remove_btn)
        toolbar.addWidget(move_up)
        toolbar.addWidget(move_down)
        toolbar.addWidget(import_btn)
        toolbar.addStretch()

        self.act_count_label = QLabel("0 activities")
        self.act_count_label.setStyleSheet("font-weight: bold; color: #7f8c8d;")
        toolbar.addWidget(self.act_count_label)
        layout.addLayout(toolbar)

        # Table
        self.act_table = QTableWidget(0, 8)
        self.act_table.setHorizontalHeaderLabels([
            "Phase", "Activity", "Depth From (m)", "Depth To (m)",
            "Duration (hrs)", "Start", "End", "Section"
        ])
        self.act_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.act_table.setColumnWidth(0, 180)
        self.act_table.setColumnWidth(2, 90)
        self.act_table.setColumnWidth(3, 90)
        self.act_table.setColumnWidth(4, 80)
        self.act_table.setColumnWidth(5, 110)
        self.act_table.setColumnWidth(6, 110)
        self.act_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.act_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.act_table.setAlternatingRowColors(True)
        self.act_table.doubleClicked.connect(self._edit_activity)
        layout.addWidget(self.act_table)

        # Summary
        self.plan_summary = QLabel("")
        self.plan_summary.setStyleSheet("font-weight: bold; color: #2c3e50; padding: 5px; background: #ecf0f1; border-radius: 3px;")
        layout.addWidget(self.plan_summary)

        return tab

    def _create_timeline_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        refresh_btn = QPushButton("🔄 Refresh Timeline")
        refresh_btn.clicked.connect(self._draw_timeline)
        layout.addWidget(refresh_btn)

        self.timeline_widget = QWidget()
        self.timeline_widget.setMinimumHeight(400)
        layout.addWidget(self.timeline_widget)

        return tab

    # ==================== Activity Management ====================

    def _add_activity_dialog(self):
        dlg = AddPlanActivityDialog(self, self.db, self.well_id, self.activities)
        if dlg.exec():
            data = dlg.get_result()
            if data:
                self.activities.append(data)
                self._refresh_table()

    def _add_phase_template(self):
        """اضافه کردن یک فاز کامل از template"""
        phases = list(DRILLING_PHASES.keys())
        phase, ok = QInputDialog.getItem(
            self, "Select Phase", "Choose a drilling phase to add:", phases, 0, False
        )
        if ok and phase:
            phase_data = DRILLING_PHASES[phase]
            color = phase_data["color"]

            for activity_name in phase_data["activities"]:
                self.activities.append({
                    "phase": phase,
                    "activity": activity_name,
                    "depth_from": 0,
                    "depth_to": 0,
                    "duration_hrs": 0,
                    "start": "",
                    "end": "",
                    "section": "",
                    "color": color,
                })
            self._refresh_table()

    def _edit_activity(self):
        row = self.act_table.currentRow()
        if 0 <= row < len(self.activities):
            dlg = AddPlanActivityDialog(
                self, self.db, self.well_id, self.activities,
                edit_data=self.activities[row]
            )
            if dlg.exec():
                data = dlg.get_result()
                if data:
                    self.activities[row] = data
                    self._refresh_table()

    def _remove_activity(self):
        row = self.act_table.currentRow()
        if 0 <= row < len(self.activities):
            self.activities.pop(row)
            self._refresh_table()

    def _move_up(self):
        row = self.act_table.currentRow()
        if row > 0:
            self.activities[row], self.activities[row-1] = self.activities[row-1], self.activities[row]
            self._refresh_table()
            self.act_table.setCurrentCell(row - 1, 0)

    def _move_down(self):
        row = self.act_table.currentRow()
        if 0 <= row < len(self.activities) - 1:
            self.activities[row], self.activities[row+1] = self.activities[row+1], self.activities[row]
            self._refresh_table()
            self.act_table.setCurrentCell(row + 1, 0)

    def _refresh_table(self):
        self.act_table.setRowCount(0)
        total_hrs = 0

        for act in self.activities:
            row = self.act_table.rowCount()
            self.act_table.insertRow(row)

            color = act.get('color', '#2c3e50')
            phase_item = QTableWidgetItem(act.get('phase', ''))
            phase_item.setForeground(QColor(color))
            phase_item.setFont(QFont("Arial", 9, QFont.Bold))
            self.act_table.setItem(row, 0, phase_item)

            self.act_table.setItem(row, 1, QTableWidgetItem(act.get('activity', '')))
            self.act_table.setItem(row, 2, QTableWidgetItem(f"{act.get('depth_from', 0):.0f}"))
            self.act_table.setItem(row, 3, QTableWidgetItem(f"{act.get('depth_to', 0):.0f}"))

            dur = act.get('duration_hrs', 0)
            total_hrs += dur
            dur_item = QTableWidgetItem(f"{dur:.1f}")
            dur_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.act_table.setItem(row, 4, dur_item)

            self.act_table.setItem(row, 5, QTableWidgetItem(str(act.get('start', ''))))
            self.act_table.setItem(row, 6, QTableWidgetItem(str(act.get('end', ''))))
            self.act_table.setItem(row, 7, QTableWidgetItem(act.get('section', '')))

            # Row color
            for col in range(self.act_table.columnCount()):
                item = self.act_table.item(row, col)
                if item and row % 2 == 0:
                    item.setBackground(QColor(color + "15"))

        total_days = total_hrs / 24
        self.act_count_label.setText(f"{len(self.activities)} activities")
        self.plan_summary.setText(
            f"📊 Total: {len(self.activities)} activities | "
            f"{total_hrs:.1f} hours ({total_days:.1f} days)"
        )

    def _draw_timeline(self):
        """رسم نمودار Gantt ساده"""
        if not self.activities:
            return

        try:
            import matplotlib
            matplotlib.use('Qt5Agg')
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

            fig, ax = plt.subplots(figsize=(10, max(4, len(self.activities) * 0.3)), facecolor='#f8f9fa')
            ax.set_facecolor('#f8f9fa')

            cum_hrs = 0
            y_pos = []
            labels = []
            colors_list = []

            for i, act in enumerate(self.activities):
                dur = act.get('duration_hrs', 0)
                if dur <= 0:
                    dur = 1

                color = act.get('color', '#3498db')
                ax.barh(i, dur, left=cum_hrs, height=0.6, color=color, alpha=0.8, edgecolor='white')

                if dur > 2:
                    ax.text(cum_hrs + dur / 2, i, f"{dur:.0f}h", ha='center', va='center', fontsize=7, color='white', fontweight='bold')

                y_pos.append(i)
                labels.append(f"{act.get('activity', '')[:25]}")
                cum_hrs += dur

            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=8)
            ax.set_xlabel("Cumulative Hours", fontsize=10)
            ax.set_title("Well Plan - Activity Timeline", fontsize=12, fontweight='bold')
            ax.invert_yaxis()
            ax.grid(axis='x', alpha=0.3)

            # Days axis on top
            ax2 = ax.twiny()
            ax2.set_xlim(ax.get_xlim()[0] / 24, ax.get_xlim()[1] / 24)
            ax2.set_xlabel("Days", fontsize=10)

            fig.tight_layout()

            canvas = FigureCanvas(fig)
            if self.timeline_widget.layout():
                while self.timeline_widget.layout().count():
                    w = self.timeline_widget.layout().takeAt(0)
                    if w.widget():
                        w.widget().deleteLater()
            else:
                self.timeline_widget.setLayout(QVBoxLayout())
                self.timeline_widget.layout().setContentsMargins(0, 0, 0, 0)
            self.timeline_widget.layout().addWidget(canvas)
            plt.close(fig)

        except Exception as e:
            logger.error(f"Timeline error: {e}")

    def _import_from_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Import Drilling Plan", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            review = PlanImportReviewDialog(filename, self)
            if review.exec() != QDialog.Accepted:
                return

            raw_activities = review.get_activities()
            if not raw_activities:
                return

            cum_depth = 0
            for act in raw_activities:
                activity = act.get("activity", "")
                iadc = act.get("iadc_code", "")
                interval = act.get("interval", 0)
                depth = act.get("depth", 0)
                hours = act.get("hours", 0)
                formation = act.get("formation", "")

                # Depth range
                depth_from = cum_depth
                depth_to = depth if depth > 0 else depth_from
                if interval > 0 and depth > 0:
                    depth_from = depth
                    depth_to = depth + interval
                elif depth > cum_depth:
                    depth_from = cum_depth
                    depth_to = depth
                if depth_to > cum_depth:
                    cum_depth = depth_to

                # Phase از IADC
                phase = self._map_iadc_to_phase(iadc)
                color = "#3498db"
                for pk, pd in DRILLING_PHASES.items():
                    if pk == phase:
                        color = pd.get("color", "#3498db")
                        break

                self.activities.append({
                    "phase": phase,
                    "activity": activity[:200],
                    "depth_from": depth_from,
                    "depth_to": depth_to,
                    "duration_hrs": hours,
                    "start": "",
                    "end": "",
                    "section": formation,
                    "color": color,
                })

            self._refresh_table()
            QMessageBox.information(
                self, "Import Complete",
                f"✅ Imported {len(raw_activities)} activities\n"
                f"📏 Max depth: {cum_depth:.0f} m"
            )

        except ImportError:
            QMessageBox.warning(
                self, "Error",
                "Install openpyxl:\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self, "Import Error", f"Failed:\n{str(e)}"
            )

    def _map_iadc_to_phase(self, iadc_code):
        if not iadc_code:
            return "10-CONTINGENCY"
        code = iadc_code.strip().lower()
        mapping = {
            "drlg": "05-PRODUCTION HOLE",
            "drilling": "05-PRODUCTION HOLE",
            "trip": "05-PRODUCTION HOLE",
            "circulation": "05-PRODUCTION HOLE",
            "cmt": "05-PRODUCTION HOLE",
            "cement": "05-PRODUCTION HOLE",
            "csg": "06-LINER SECTION",
            "casing": "06-LINER SECTION",
            "liner": "06-LINER SECTION",
            "log": "05-PRODUCTION HOLE",
            "bop test": "03-SURFACE HOLE",
            "bop": "03-SURFACE HOLE",
            "milling": "10-CONTINGENCY",
            "mill": "10-CONTINGENCY",
            "completion": "07-COMPLETION",
            "perforation": "07-COMPLETION",
            "dry test": "08-TESTING",
            "test": "08-TESTING",
            "cut & slip": "09-RIG DOWN & MOVE",
            "rig service": "09-RIG DOWN & MOVE",
        }
        for key, phase in mapping.items():
            if key in code:
                return phase
        return "05-PRODUCTION HOLE"

    # ==================== Load/Save ====================

    def load_well_info(self):
        try:
            well = self.db.get_well_by_id(self.well_id)
            if well:
                self.header_label.setText(f"📋 Well Plan - {well.get('name', '')}")
                self.final_depth.setValue(well.get('target_depth', 0) or 0)
                if well.get('spud_date'):
                    try:
                        sd = well['spud_date']
                        if isinstance(sd, str):
                            from datetime import datetime
                            sd = datetime.strptime(sd, "%Y-%m-%d").date()
                        self.spud_date.setDate(QDate(sd.year, sd.month, sd.day))
                    except:
                        pass
        except Exception as e:
            logger.error(f"Load well info error: {e}")

    def save_plan(self):
        session = self.db.create_session()
        try:
            plan = WellPlan(
                well_id=self.well_id,
                plan_name=self.plan_name.text() or f"Plan {datetime.now().strftime('%Y-%m-%d')}",
                plan_version=self.plan_version.text() or "1.0",
                planned_spud_date=self.spud_date.date().toPython(),
                planned_finish_date=self.finish_date.date().toPython(),
                planned_total_days=self.spud_date.date().daysTo(self.finish_date.date()),
                planned_final_depth=self.final_depth.value(),
                description=self.description.toPlainText(),
                is_active=True
            )
            session.add(plan)
            session.flush()

            for act in self.activities:
                # Parse dates
                start_dt = None
                end_dt = None
                try:
                    if act.get('start'):
                        start_dt = datetime.strptime(str(act['start']), "%Y-%m-%d %H:%M")
                except:
                    start_dt = datetime.now()

                try:
                    if act.get('end'):
                        end_dt = datetime.strptime(str(act['end']), "%Y-%m-%d %H:%M")
                except:
                    end_dt = datetime.now() + timedelta(hours=act.get('duration_hrs', 0))

                if not start_dt:
                    start_dt = datetime.now()
                if not end_dt:
                    end_dt = start_dt + timedelta(hours=act.get('duration_hrs', 0))

                pa = PlannedActivity(
                    well_id=self.well_id,
                    plan_id=plan.id,
                    activity_name=act.get('activity', ''),
                    phase_code=act.get('phase', ''),
                    planned_start=start_dt,
                    planned_end=end_dt,
                    planned_duration_hours=act.get('duration_hrs', 0),
                    planned_depth_from=act.get('depth_from', 0),
                    planned_depth_to=act.get('depth_to', 0),
                )
                session.add(pa)

            session.commit()
            self.plan_id = plan.id
            QMessageBox.information(self, "Success", f"Plan saved! ({len(self.activities)} activities)")
            self.accept()

        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Save failed: {str(e)}")
            logger.error(f"Save plan error: {e}")
        finally:
            session.close()


class AddPlanActivityDialog(QDialog):
    """دیالوگ اضافه/ویرایش فعالیت برنامه"""

    def __init__(self, parent, db_manager, well_id, existing_activities, edit_data=None):
        super().__init__(parent)
        self.db = db_manager
        self.well_id = well_id
        self.existing = existing_activities
        self.edit_data = edit_data
        self.result = None

        self.setWindowTitle("📅 Plan Activity" if not edit_data else "📅 Edit Activity")
        self.setMinimumWidth(500)
        self.init_ui()
        if edit_data:
            self._load(edit_data)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # Phase
        g1 = QGroupBox("📋 Phase & Activity")
        f1 = QFormLayout(g1)

        self.phase_combo = QComboBox()
        for phase_name in DRILLING_PHASES:
            color = DRILLING_PHASES[phase_name]["color"]
            self.phase_combo.addItem(f"🔵 {phase_name}", phase_name)
        self.phase_combo.currentIndexChanged.connect(self._on_phase_changed)
        f1.addRow("Phase:", self.phase_combo)

        self.activity_combo = QComboBox()
        self.activity_combo.setEditable(True)
        f1.addRow("Activity:", self.activity_combo)

        layout.addWidget(g1)

        # Depth & Duration
        g2 = QGroupBox("📏 Depth & Duration")
        f2 = QFormLayout(g2)

        self.depth_from = QDoubleSpinBox()
        self.depth_from.setRange(0, 20000)
        self.depth_from.setSuffix(" m")
        f2.addRow("Depth From:", self.depth_from)

        self.depth_to = QDoubleSpinBox()
        self.depth_to.setRange(0, 20000)
        self.depth_to.setSuffix(" m")
        f2.addRow("Depth To:", self.depth_to)

        self.duration = QDoubleSpinBox()
        self.duration.setRange(0, 10000)
        self.duration.setDecimals(1)
        self.duration.setSuffix(" hrs")
        f2.addRow("Duration:", self.duration)

        self.days_label = QLabel("= 0.0 days")
        self.days_label.setStyleSheet("color: #3498db; font-weight: bold;")
        f2.addRow("", self.days_label)
        self.duration.valueChanged.connect(
            lambda v: self.days_label.setText(f"= {v/24:.1f} days")
        )

        layout.addWidget(g2)

        # Time
        g3 = QGroupBox("📅 Schedule (optional)")
        f3 = QFormLayout(g3)

        self.start_dt = QDateTimeEdit(QDateTime.currentDateTime())
        self.start_dt.setCalendarPopup(True)
        self.start_dt.setDisplayFormat("yyyy-MM-dd HH:mm")
        f3.addRow("Start:", self.start_dt)

        self.end_dt = QDateTimeEdit(QDateTime.currentDateTime().addDays(1))
        self.end_dt.setCalendarPopup(True)
        self.end_dt.setDisplayFormat("yyyy-MM-dd HH:mm")
        f3.addRow("End:", self.end_dt)

        layout.addWidget(g3)

        # Section
        g4 = QGroupBox("📊 Section")
        f4 = QFormLayout(g4)
        self.section_combo = QComboBox()
        self.section_combo.addItem("-- None --", None)
        if self.db and self.well_id:
            sections = self.db.get_sections_by_well(self.well_id)
            for s in sections:
                self.section_combo.addItem(s['name'], s['id'])
        f4.addRow("Section:", self.section_combo)
        layout.addWidget(g4)

        # Buttons
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("✅ Add" if not self.edit_data else "✅ Update")
        save_btn.setStyleSheet("background: #27ae60; color: white; font-weight: bold; padding: 8px 20px; border-radius: 4px; border: none;")
        save_btn.clicked.connect(self._save)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self._on_phase_changed()

    def _on_phase_changed(self):
        phase_key = self.phase_combo.currentData()
        self.activity_combo.clear()
        if phase_key and phase_key in DRILLING_PHASES:
            self.activity_combo.addItems(DRILLING_PHASES[phase_key]["activities"])

    def _load(self, data):
        # Phase
        phase = data.get('phase', '')
        for i in range(self.phase_combo.count()):
            if self.phase_combo.itemData(i) == phase:
                self.phase_combo.setCurrentIndex(i)
                break
        # Activity
        act = data.get('activity', '')
        idx = self.activity_combo.findText(act)
        if idx >= 0:
            self.activity_combo.setCurrentIndex(idx)
        else:
            self.activity_combo.setCurrentText(act)

        self.depth_from.setValue(data.get('depth_from', 0))
        self.depth_to.setValue(data.get('depth_to', 0))
        self.duration.setValue(data.get('duration_hrs', 0))

    def _save(self):
        phase_key = self.phase_combo.currentData() or ""
        color = DRILLING_PHASES.get(phase_key, {}).get("color", "#3498db")

        self.result = {
            "phase": phase_key,
            "activity": self.activity_combo.currentText(),
            "depth_from": self.depth_from.value(),
            "depth_to": self.depth_to.value(),
            "duration_hrs": self.duration.value(),
            "start": self.start_dt.dateTime().toString("yyyy-MM-dd HH:mm"),
            "end": self.end_dt.dateTime().toString("yyyy-MM-dd HH:mm"),
            "section": self.section_combo.currentText() if self.section_combo.currentData() else "",
            "color": color,
        }
        self.accept()

    def get_result(self):
        return self.result